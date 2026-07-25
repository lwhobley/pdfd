"""PDF → Excel/CSV table extraction tool.

Uses pdfplumber for table detection.
Exports each table to a sheet in an .xlsx file, or concatenates to .csv.
"""
from __future__ import annotations
import csv
import os
from typing import Any

from pdf_forge.tools.base import BaseTool, ToolMeta
from pdf_forge.workers.job_model import Job, JobResult


class PDFToExcelJob(Job):
    """Extract tables from PDF and write to Excel (.xlsx) or CSV."""

    def __init__(
        self,
        input_path: str,
        output_path: str,
        pages: list[int] | None = None,
        fmt: str = "xlsx",
    ) -> None:
        super().__init__("pdf_to_excel", [input_path])
        self.output_path = output_path
        self.pages = pages
        self.fmt = fmt.lower()

    def execute(self) -> JobResult:
        import pdfplumber

        self.log(f"Extracting tables → {self.fmt.upper()}")

        with pdfplumber.open(self.input_paths[0]) as pdf:
            target = (
                [pdf.pages[i] for i in self.pages]
                if self.pages is not None
                else pdf.pages
            )

            all_tables: list[tuple[int, int, list[list]]] = []
            for pg_idx, page in enumerate(target):
                if self.cancel_flag:
                    raise Exception("Cancelled")
                tables = page.extract_tables()
                for tbl_idx, table in enumerate(tables):
                    if table:
                        all_tables.append((pg_idx + 1, tbl_idx + 1, table))
                self.report_progress(int((pg_idx + 1) / len(target) * 85))

        self.log(f"Found {len(all_tables)} table(s)")

        if self.fmt == "csv":
            self._write_csv(all_tables)
        else:
            self._write_xlsx(all_tables)

        self.report_progress(100)
        return JobResult(
            output_paths=[self.output_path],
            metadata={"table_count": len(all_tables)},
        )

    def _write_csv(self, tables: list) -> None:
        with open(self.output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for pg_num, tbl_num, rows in tables:
                writer.writerow([f"--- Page {pg_num}, Table {tbl_num} ---"])
                for row in rows:
                    writer.writerow([c or "" for c in row])
                writer.writerow([])

    def _write_xlsx(self, tables: list) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        if not tables:
            wb.active.title = "No Tables Found"
            wb.save(self.output_path)
            return

        wb.remove(wb.active)
        for pg_num, tbl_num, rows in tables:
            sheet_name = f"P{pg_num}_T{tbl_num}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, cell in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=cell or "")
                    if r_idx == 1:
                        ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)

        wb.save(self.output_path)


class PDFToExcelTool(BaseTool):
    meta = ToolMeta(
        tool_id="pdf_to_excel",
        name="PDF → Excel/CSV",
        description="Extract tables from a PDF to an Excel workbook or CSV file.",
        category="convert",
        icon="pdf_to_excel",
    )

    def create_job(self, params: dict[str, Any]) -> PDFToExcelJob:
        return PDFToExcelJob(
            input_path=params["input_path"],
            output_path=params["output_path"],
            pages=params.get("pages"),
            fmt=params.get("fmt", "xlsx"),
        )
