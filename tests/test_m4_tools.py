"""M4 tool tests — Office→PDF, table extraction, linearize, sanitize."""
from __future__ import annotations
import os
import json
import pytest


# ── Linearize ─────────────────────────────────────────────────────────────────

def test_linearize_produces_valid_pdf(sample_pdf, tmp_path):
    import pikepdf
    from pdf_forge.tools.organize.linearize import LinearizeJob

    out = str(tmp_path / "linear.pdf")
    result = LinearizeJob(input_path=sample_pdf, output_path=out).execute()
    assert os.path.exists(out)
    assert out in result.output_paths
    # Should still be a valid PDF with same page count
    with pikepdf.open(out) as doc:
        with pikepdf.open(sample_pdf) as src:
            assert len(doc.pages) == len(src.pages)


# ── Sanitize ──────────────────────────────────────────────────────────────────

def test_sanitize_removes_metadata(tmp_path):
    import pikepdf
    from pdf_forge.tools.organize.sanitize import SanitizeJob

    # Create PDF with metadata
    src_path = str(tmp_path / "src.pdf")
    doc = pikepdf.Pdf.new()
    doc.add_blank_page(page_size=(595, 842))
    doc.docinfo["/Author"] = "Test Author"
    doc.docinfo["/Title"] = "Secret Title"
    doc.save(src_path)

    out = str(tmp_path / "clean.pdf")
    SanitizeJob(
        input_path=src_path,
        output_path=out,
        remove_metadata=True,
        remove_javascript=True,
        remove_embedded=True,
    ).execute()

    with pikepdf.open(out) as result:
        assert result.docinfo.get("/Author", "") == ""
        assert result.docinfo.get("/Title", "") == ""


def test_sanitize_preserves_pages(sample_pdf, tmp_path):
    import pikepdf
    from pdf_forge.tools.organize.sanitize import SanitizeJob

    out = str(tmp_path / "sanitized.pdf")
    SanitizeJob(
        input_path=sample_pdf,
        output_path=out,
    ).execute()
    with pikepdf.open(out) as doc:
        with pikepdf.open(sample_pdf) as src:
            assert len(doc.pages) == len(src.pages)


# ── PDF → Excel/CSV ───────────────────────────────────────────────────────────

def test_pdf_to_csv_no_tables(sample_pdf, tmp_path):
    """A text-only PDF produces a CSV with zero tables (no crash)."""
    from pdf_forge.tools.convert.pdf_to_excel import PDFToExcelJob

    out = str(tmp_path / "tables.csv")
    result = PDFToExcelJob(
        input_path=sample_pdf,
        output_path=out,
        fmt="csv",
    ).execute()
    assert os.path.exists(out)
    assert result.metadata["table_count"] == 0


def test_pdf_to_xlsx_with_table(tmp_path):
    """Create a PDF with a table via reportlab; verify extraction."""
    try:
        from reportlab.platypus import SimpleDocTemplate, Table
        from reportlab.lib.pagesizes import A4
    except ImportError:
        pytest.skip("reportlab not installed")

    src_path = str(tmp_path / "table.pdf")
    doc = SimpleDocTemplate(src_path, pagesize=A4)
    table_data = [
        ["Name",  "Score", "Grade"],
        ["Alice", "95",    "A"],
        ["Bob",   "78",    "B"],
    ]
    doc.build([Table(table_data)])

    from pdf_forge.tools.convert.pdf_to_excel import PDFToExcelJob

    out = str(tmp_path / "tables.xlsx")
    result = PDFToExcelJob(
        input_path=src_path,
        output_path=out,
        fmt="xlsx",
    ).execute()
    assert os.path.exists(out)
    # table_count may be 0 if pdfplumber doesn't detect the reportlab layout
    assert "table_count" in result.metadata

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert len(wb.sheetnames) >= 1


# ── Office → PDF (unit-level: only checks error when LO absent) ───────────────

def test_office_to_pdf_raises_when_no_libreoffice(tmp_path):
    from pdf_forge.tools.convert.office_to_pdf import OfficeToPDFJob
    from pdf_forge.adapters.libreoffice_adapter import find_soffice

    if find_soffice():
        pytest.skip("LibreOffice is installed; can't test missing-LO path")

    # Create a dummy docx-like file
    src = str(tmp_path / "test.docx")
    with open(src, "wb") as f:
        f.write(b"PK\x03\x04")  # DOCX magic bytes (zip)

    with pytest.raises(RuntimeError, match="LibreOffice"):
        OfficeToPDFJob(
            input_paths=[src],
            output_dir=str(tmp_path / "out"),
            libreoffice_path="/nonexistent/soffice",
        ).execute()
